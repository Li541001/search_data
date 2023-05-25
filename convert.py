from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
import json
import tkinter
from tkinter import filedialog


def main():
  new_object = {}
  xlsx = load_workbook(path.name)
  xlsxs = xlsx.active
  #連接excel

  response = requests.get('https://api.jsonstorage.net/v1/json/4eed805d-59cf-4ff8-ad3b-faf98c316270/c0e5a20d-e108-45d2-9097-19ef727e7af4')
  # print(f'oldData: {response.text}')
  new_object = json.loads(response.text)
  #取得database

  i = 1
  while True:
      i += 1
      value = []
      title = []
      id = str(xlsxs['A'+str(i)].value)
      id_object={}
      new_object.update({id:id_object})
      if(id == "None"):
        break
      id_object = new_object[id]

      j = 1
      while True:
          j += 1
          char = get_column_letter(j)
          title = xlsxs[str(char)+'1'].value
          value = str(xlsxs[str(char)+str(i)].value)

          if(value == "None"):
            break
          id_object.update({title:value})
      new_object.update({id:id_object})
  #迴圈取得成績並轉換成正確形式

  json_object = json.dumps(new_object)  #轉成json
  print(f'giveData: {json_object}')

  r = requests.put('https://api.jsonstorage.net/v1/json/4eed805d-59cf-4ff8-ad3b-faf98c316270/c0e5a20d-e108-45d2-9097-19ef727e7af4?apiKey=a2ddf8a1-e69b-4a56-84be-ed30fa49a6ad',data=json_object,headers = {"Content-type": "application/json"})
  print(r)
  #put到database

  response = requests.get('https://api.jsonstorage.net/v1/json/4eed805d-59cf-4ff8-ad3b-faf98c316270/c0e5a20d-e108-45d2-9097-19ef727e7af4')
  print(f'newData: {response.text}')
  print()
  print('------------------------------------finish----------------------------------------')
  #再次讀取database

def fileStoryLocation():
    global path
    pathset = filedialog.askopenfile()
    file_path.set(pathset.name)
    path = pathset
    #讀取電腦檔案

if __name__ == '__main__':
    top = tkinter.Tk()
    top.title('選擇檔案')
    label = tkinter.Label(top,text="儲存位置: ")
    label.grid(row=0,column = 0)
    file_path = tkinter.StringVar()
    pathInput = tkinter.Entry(top,textvariable=file_path)
    pathInput.grid(row=0,column = 1)
    submit = tkinter.Button(top,command=fileStoryLocation,text="檔案位置")
    submit.grid(row = 0,column = 2)
    submit = tkinter.Button(top,command=main,text="確定")
    submit.grid(row = 1,column = 1)
    top.mainloop()
    #ui介面


'''
{
  0907119:{
    chinese:76,
    english:81.5,
    math:88
  }
}

'''
